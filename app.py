# app.py
import os, sqlite3, hashlib, time
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from functools import wraps
from dsa_shared import DSALib, DSALibError
from fpdf import FPDF
from openpyxl import Workbook
import matplotlib.pyplot as plt
from io import BytesIO

# Config
APP_SECRET = "change_this_to_a_random_secret_in_production"
DB_FILE = "students.db"
STATIC_CHART_DIR = os.path.join("static", "charts")
os.makedirs(STATIC_CHART_DIR, exist_ok=True)

app = Flask(__name__, static_folder="static")
app.secret_key = APP_SECRET

# ---------- DB helpers ----------
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            roll INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            marks INTEGER NOT NULL
        )""")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT CHECK(role IN ('admin','teacher','student')) NOT NULL
        )""")
    conn.commit()
    conn.close()

def hash_password(p): return hashlib.sha256(p.encode()).hexdigest()

# ---------- auth decorators ----------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def wrapper(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'role' not in session or session['role'] not in roles:
                flash("You don't have permission to access that page.", "danger")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return wrapper

# ---------- small helpers ----------
def current_user():
    return session.get('username'), session.get('role')

def build_csv_string(rows):
    return "".join(f"{r['roll']},{r['name']},{r['marks']}\n" for r in rows)

# ---------- routes ----------
@app.route("/")
def index():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT password_hash, role FROM users WHERE username=?", (username,))
        row = cur.fetchone()
        conn.close()
        if row and row['password_hash'] == hash_password(password):
            session['username'] = username
            session['role'] = row['role']
            flash(f"Welcome {username}!", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid credentials", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out", "info")
    return redirect(url_for('login'))

@app.route("/dashboard")
@login_required
def dashboard():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll, name, marks FROM students ORDER BY roll")
    else:
        # student sees only own record (username like student1)
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll, name, marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall()
    conn.close()
    return render_template("dashboard.html", rows=rows, user=username, role=role)

# ---------- CRUD ----------
@app.route("/students", methods=["GET","POST"])
@login_required
def students():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if request.method == "POST":
        if role not in ("admin","teacher"):
            flash("Unauthorized", "danger"); return redirect(url_for('students'))
        # Add new
        try:
            roll = int(request.form['roll'])
            name = request.form['name'].strip()
            marks = int(request.form['marks'])
            cur.execute("INSERT INTO students (roll, name, marks) VALUES (?, ?, ?)", (roll, name, marks))
            conn.commit()
            flash("Student added", "success")
        except sqlite3.IntegrityError:
            flash("Roll already exists", "danger")
        except Exception as e:
            flash(f"Invalid input: {e}", "danger")
        return redirect(url_for('students'))
    # GET
    if role in ("admin","teacher"):
        cur.execute("SELECT roll, name, marks FROM students ORDER BY roll")
    else:
        # student
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll, name, marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall()
    conn.close()
    return render_template("students.html", rows=rows, user=username, role=role)

@app.route("/students/edit/<int:roll>", methods=["GET","POST"])
@login_required
@role_required("admin","teacher")
def edit_student(roll):
    conn = get_db(); cur = conn.cursor()
    if request.method == "POST":
        name = request.form['name'].strip(); marks = int(request.form['marks'])
        cur.execute("UPDATE students SET name=?, marks=? WHERE roll=?", (name, marks, roll))
        conn.commit(); conn.close()
        flash("Updated", "success")
        return redirect(url_for('students'))
    cur.execute("SELECT roll, name, marks FROM students WHERE roll=?", (roll,))
    row = cur.fetchone(); conn.close()
    if not row: flash("Not found", "danger"); return redirect(url_for('students'))
    return render_template("edit_student.html", r=row)

@app.route("/students/delete/<int:roll>", methods=["POST"])
@login_required
@role_required("admin")
def delete_student(roll):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM students WHERE roll=?", (roll,))
    conn.commit(); conn.close()
    flash("Deleted", "info")
    return redirect(url_for('students'))

# ---------- DSA endpoints (use shared lib in-memory) ----------
@app.route("/dsa/sort/<by>")
@login_required
def dsa_sort(by):
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll,name,marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll,name,marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    csv_str = build_csv_string(rows)
    dsa = DSALib()
    try:
        dsa.load_from_string(csv_str)
        if by == "roll":
            dsa.sort_by_roll()
        else:
            dsa.sort_by_name()
        out = dsa.export_to_string()
        dsa.free()
        # turn into list for template
        lines = [line for line in out.splitlines() if line.strip()]
        entries = [tuple(line.split(",")) for line in lines]
        return render_template("dsa_result.html", entries=entries, user=username, role=role, action=f"sort by {by}")
    except DSALibError as e:
        flash(f"DSA error: {e}", "danger")
        return redirect(url_for('dashboard'))

@app.route("/dsa/search", methods=["POST"])
@login_required
def dsa_search():
    username, role = current_user()
    key = request.form.get("roll_key")
    if not key or not key.isdigit():
        flash("Invalid roll", "danger"); return redirect(url_for('dashboard'))
    key = int(key)
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll,name,marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll,name,marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    csv_str = build_csv_string(rows)
    dsa = DSALib()
    try:
        dsa.load_from_string(csv_str)
        res = dsa.search_roll(key)
        dsa.free()
        if res:
            flash(f"Found: {res}", "success")
        else:
            flash("Not found", "info")
    except DSALibError as e:
        flash(f"DSA error: {e}", "danger")
    return redirect(url_for('dashboard'))

@app.route("/dsa/stats")
@login_required
def dsa_stats():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll,name,marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll,name,marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    csv_str = build_csv_string(rows)
    dsa = DSALib()
    try:
        dsa.load_from_string(csv_str)
        st = dsa.stats()
        dsa.free()
        return render_template("dsa_stats.html", stats=st, user=username, role=role)
    except DSALibError as e:
        flash(f"DSA error: {e}", "danger")
        return redirect(url_for('dashboard'))

# ---------- Exports ----------
@app.route("/export/pdf")
@login_required
def export_pdf():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll,name,marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll,name,marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    if not rows:
        flash("No data", "danger"); return redirect(url_for('dashboard'))

    pdf = FPDF(); pdf.add_page(); pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Student Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"Generated by: {username} ({role})", ln=True)
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(30,8,"Roll",1); pdf.cell(90,8,"Name",1); pdf.cell(30,8,"Marks",1); pdf.ln()
    pdf.set_font("Helvetica","",12)
    for r in rows:
        pdf.cell(30,8,str(r['roll']),1); pdf.cell(90,8,str(r['name']),1); pdf.cell(30,8,str(r['marks']),1); pdf.ln()
    # stats
    csv_str = build_csv_string(rows)
    dsa = DSALib(); dsa.load_from_string(csv_str); st = dsa.stats(); dsa.free()
    pdf.ln(4); pdf.cell(0,8,f"Count: {st['count']}  Avg: {st['avg']:.2f}  Min: {st['min']}  Max: {st['max']}", ln=True)
    # send as file
    buf = BytesIO(); pdf.output(buf); buf.seek(0)
    return send_file(buf, download_name="students_report.pdf", as_attachment=True, mimetype="application/pdf")

@app.route("/export/excel")
@login_required
def export_excel():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT roll,name,marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT roll,name,marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    wb = Workbook(); ws = wb.active; ws.append(["Roll","Name","Marks"])
    for r in rows: ws.append([r['roll'], r['name'], r['marks']])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="students.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- Charts ----------
@app.route("/chart")
@login_required
def chart():
    username, role = current_user()
    conn = get_db(); cur = conn.cursor()
    if role in ("admin","teacher"):
        cur.execute("SELECT name, marks FROM students ORDER BY roll")
    else:
        roll = ''.join([c for c in username if c.isdigit()]) or '1'
        cur.execute("SELECT name, marks FROM students WHERE roll=?", (roll,))
    rows = cur.fetchall(); conn.close()
    if not rows:
        flash("No data", "danger"); return redirect(url_for('dashboard'))
    names = [r['name'] for r in rows]
    marks = [r['marks'] for r in rows]
    # plot
    plt.figure(figsize=(8,4))
    plt.bar(names, marks)
    plt.title("Marks Distribution")
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    fname = f"chart_{int(time.time())}.png"
    fpath = os.path.join(STATIC_CHART_DIR, fname)
    plt.savefig(fpath)
    plt.close()
    return render_template("chart.html", chart_url=url_for('static', filename=f"charts/{fname}"), user=username, role=role)

# ---------- utilities ----------
@app.route("/init_sample")
def init_sample():
    # convenience endpoint to populate sample data
    from sample_data import main as populate_main
    populate_main()
    flash("Sample data populated", "success")
    return redirect(url_for('dashboard'))

# ---------- startup ----------
if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
